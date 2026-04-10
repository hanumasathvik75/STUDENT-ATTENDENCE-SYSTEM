import csv
import os
from pathlib import Path
from datetime import datetime


class DFA:
    def __init__(self):
        self.states = {"q0", "q1", "q2", "q3", "q4", "q5", "q6"}
        self.start_state = "q0"
        self.accept_states = {"q6"}
        self.current_state = self.start_state
    
    def transition(self, state, symbol):
        if state == "q0":
            return "q1" if symbol == '2' else None
        elif state == "q1":
            return "q2" if symbol == '4' else None
        elif state == "q2":
            return "q3" if symbol == 'B' else None
        elif state == "q3":
            return "q4" if (symbol.isalpha() and symbol.isupper()) else None
        elif state == "q4":
            return "q5" if (symbol.isalpha() and symbol.isupper()) else None
        elif state == "q5":
            return "q6" if symbol in ['0', '1', '2', '3', '4'] else None
        else:
            return None
    
    def validate(self, reg_number):
        if not isinstance(reg_number, str) or len(reg_number) != 6:
            return False
        
        self.current_state = self.start_state
        
        for symbol in reg_number:
            self.current_state = self.transition(self.current_state, symbol)
            if self.current_state is None:
                return False
        
        return self.current_state in self.accept_states


class StudentDatabase:
    def __init__(self, folder="student_data"):
        self.folder = folder
        self.students_file = os.path.join(folder, "students.csv")
        self.attendance_file = os.path.join(folder, "attendance.csv")
        
        Path(folder).mkdir(exist_ok=True)
        self._init_files()
    
    def _init_files(self):
        if not os.path.exists(self.students_file):
            with open(self.students_file, 'w', newline='') as f:
                csv.writer(f).writerow(['RegNo', 'Name', 'Date'])
        
        if not os.path.exists(self.attendance_file):
            with open(self.attendance_file, 'w', newline='') as f:
                csv.writer(f).writerow(['Date', 'RegNo', 'Name', 'Status', 'Time'])
    
    def add_student(self, reg_no, name):
        dfa = DFA()
        if not dfa.validate(reg_no):
            return False, "Invalid format: 24B[A-Z][A-Z][0-4]"
        
        with open(self.students_file, 'r') as f:
            students = csv.reader(f)
            for row in students:
                if row and row[0] == reg_no:
                    return False, "Student already exists"
        
        with open(self.students_file, 'a', newline='') as f:
            csv.writer(f).writerow([reg_no, name, datetime.now().strftime("%Y-%m-%d %H:%M")])
        
        return True, f"Student {reg_no} added successfully"
    
    def mark_attendance(self, reg_no, status):
        with open(self.students_file, 'r') as f:
            students = list(csv.reader(f))
            student_name = None
            for row in students[1:]:
                if row and row[0] == reg_no:
                    student_name = row[1]
                    break
        
        if not student_name:
            return False, "Student not found"
        
        with open(self.attendance_file, 'a', newline='') as f:
            csv.writer(f).writerow([
                datetime.now().strftime("%Y-%m-%d"),
                reg_no,
                student_name,
                status,
                datetime.now().strftime("%H:%M:%S")
            ])
        
        return True, f"Marked {status}"
    
    def get_students(self):
        try:
            with open(self.students_file, 'r') as f:
                reader = csv.reader(f)
                next(reader)
                return [row for row in reader if row]
        except:
            return []
    
    def get_attendance(self):
        try:
            with open(self.attendance_file, 'r') as f:
                reader = csv.reader(f)
                next(reader)
                return [row for row in reader if row]
        except:
            return []
    
    def export_excel(self, filename):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            
            ws1 = wb.active
            ws1.title = "Students"
            with open(self.students_file) as f:
                for r_idx, row in enumerate(csv.reader(f), 1):
                    for c_idx, val in enumerate(row, 1):
                        ws1.cell(r_idx, c_idx, val)
            
            for row in ws1.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
                    cell.font = Font(bold=True, color="ffffff")
            
            ws1.column_dimensions['A'].width = 12
            ws1.column_dimensions['B'].width = 20
            ws1.column_dimensions['C'].width = 18
            
            ws2 = wb.create_sheet("Attendance")
            with open(self.attendance_file) as f:
                for r_idx, row in enumerate(csv.reader(f), 1):
                    for c_idx, val in enumerate(row, 1):
                        ws2.cell(r_idx, c_idx, val)
            
            for row in ws2.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
                    cell.font = Font(bold=True, color="ffffff")
            
            ws2.column_dimensions['A'].width = 12
            ws2.column_dimensions['B'].width = 12
            ws2.column_dimensions['C'].width = 20
            ws2.column_dimensions['D'].width = 12
            ws2.column_dimensions['E'].width = 12
            
            wb.save(filename)
            return True, f"Exported to {filename}"
        except Exception as e:
            return False, str(e)
    
    def clear_all_data(self):
        with open(self.students_file, 'w', newline='') as f:
            csv.writer(f).writerow(['RegNo', 'Name', 'Date'])
        with open(self.attendance_file, 'w', newline='') as f:
            csv.writer(f).writerow(['Date', 'RegNo', 'Name', 'Status', 'Time'])
    
    def delete_student(self, reg_no):
        students = []
        with open(self.students_file, 'r') as f:
            reader = csv.reader(f)
            header = next(reader)
            students = [row for row in reader if row and row[0] != reg_no]
        
        with open(self.students_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['RegNo', 'Name', 'Date'])
            writer.writerows(students)
        
        return True, f"Student {reg_no} deleted"
